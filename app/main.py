from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

from .db import ensure_default_manuscript, get_conn, init_db

REQUIRED_IMPORT_COLUMNS = [
    "reviewer_id",
    "line_number",
    "verbatim_comment",
    "comment_category",
]
VALID_CATEGORIES = {"editorial", "major", "minor"}
REVIEWER_RE = re.compile(r"^R(\d+)$", re.IGNORECASE)
EDITOR_RE = re.compile(r"^E(\d+)$", re.IGNORECASE)

app = FastAPI(title="Reviewer Ticket Dashboard")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_line_number(value: Any) -> tuple[str, int]:
    raw = "" if value is None else str(value).strip()
    if not raw:
        return "", 2_147_483_647
    if raw.isdigit():
        num = int(raw)
        return raw, num
    match = re.search(r"(\d+)", raw)
    if match:
        return raw, int(match.group(1))
    return raw, 2_147_483_647


def parse_reviewer_sort(reviewer_id: str) -> tuple[int, int]:
    value = (reviewer_id or "").strip()
    reviewer_match = REVIEWER_RE.match(value)
    if reviewer_match:
        return 0, int(reviewer_match.group(1))
    editor_match = EDITOR_RE.match(value)
    if editor_match:
        return 2, int(editor_match.group(1))
    fallback_num_match = re.search(r"(\d+)", value)
    fallback_num = int(fallback_num_match.group(1)) if fallback_num_match else 2_147_483_647
    return 1, fallback_num


def ensure_manuscript_exists(manuscript_id: int) -> None:
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM manuscripts WHERE id = ?", (manuscript_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Manuscript not found")


def normalize_ticket_row(row: dict[str, Any]) -> dict[str, Any]:
    reviewer_id = str(row.get("reviewer_id", "")).strip()
    if not reviewer_id:
        raise ValueError("reviewer_id is required")

    line_display, line_sort = parse_line_number(row.get("line_number", ""))
    if not line_display:
        raise ValueError("line_number is required")

    verbatim_comment = str(row.get("verbatim_comment", "")).strip()
    if not verbatim_comment:
        raise ValueError("verbatim_comment is required")

    category = str(row.get("comment_category", "")).strip().lower()
    if category not in VALID_CATEGORIES:
        raise ValueError("comment_category must be one of editorial, major, minor")

    reviewer_group_sort, reviewer_num_sort = parse_reviewer_sort(reviewer_id)

    return {
        "reviewer_id": reviewer_id,
        "reviewer_group_sort": reviewer_group_sort,
        "reviewer_num_sort": reviewer_num_sort,
        "line_number_display": line_display,
        "line_number_sort": line_sort,
        "verbatim_comment": verbatim_comment,
        "comment_category": category,
    }


def query_tickets(
    manuscript_id: int,
    search: str | None = None,
    reviewer_id: str | None = None,
    comment_category: str | None = None,
    status: str | None = None,
) -> list[dict[str, Any]]:
    sql = """
        SELECT
            id,
            manuscript_id,
            reviewer_id,
            reviewer_group_sort,
            reviewer_num_sort,
            line_number_display,
            line_number_sort,
            verbatim_comment,
            comment_category,
            response_text,
            status,
            created_at,
            updated_at
        FROM tickets
        WHERE manuscript_id = ?
    """
    params: list[Any] = [manuscript_id]

    if search:
        sql += " AND (LOWER(verbatim_comment) LIKE ? OR LOWER(response_text) LIKE ?)"
        like = f"%{search.strip().lower()}%"
        params.extend([like, like])
    if reviewer_id:
        sql += " AND reviewer_id = ?"
        params.append(reviewer_id)
    if comment_category:
        sql += " AND comment_category = ?"
        params.append(comment_category)
    if status and status in {"OPEN", "COMPLETED"}:
        sql += " AND status = ?"
        params.append(status)

    sql += """
        ORDER BY
            reviewer_group_sort ASC,
            reviewer_num_sort ASC,
            line_number_sort ASC,
            created_at ASC,
            id ASC
    """

    with get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [dict(row) for row in rows]


def get_filter_values(manuscript_id: int) -> dict[str, list[str]]:
    with get_conn() as conn:
        reviewer_rows = conn.execute(
            """
            SELECT DISTINCT reviewer_id, reviewer_group_sort, reviewer_num_sort
            FROM tickets
            WHERE manuscript_id = ?
            ORDER BY reviewer_group_sort, reviewer_num_sort, reviewer_id
            """,
            (manuscript_id,),
        ).fetchall()
    return {"reviewer_ids": [row["reviewer_id"] for row in reviewer_rows], "categories": sorted(VALID_CATEGORIES)}


def parse_csv_upload(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail="CSV file is empty")

    header = [h.strip() for h in rows[0]]
    if header != REQUIRED_IMPORT_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=f"CSV columns must be exactly: {', '.join(REQUIRED_IMPORT_COLUMNS)}",
        )

    parsed: list[dict[str, Any]] = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(cell.strip() for cell in row):
            continue
        if len(row) != 4:
            raise HTTPException(status_code=400, detail=f"Row {i}: expected 4 columns")
        parsed.append(
            {
                "reviewer_id": row[0],
                "line_number": row[1],
                "verbatim_comment": row[2],
                "comment_category": row[3],
            }
        )
    return parsed


def parse_xlsx_upload(data: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="XLSX file is empty")
    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if header != REQUIRED_IMPORT_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=f"XLSX columns must be exactly: {', '.join(REQUIRED_IMPORT_COLUMNS)}",
        )

    parsed: list[dict[str, Any]] = []
    for i, row in enumerate(rows[1:], start=2):
        values = list(row)
        if not any(value not in (None, "") for value in values):
            continue
        if len(values) < 4:
            raise HTTPException(status_code=400, detail=f"Row {i}: expected 4 columns")
        parsed.append(
            {
                "reviewer_id": "" if values[0] is None else str(values[0]),
                "line_number": "" if values[1] is None else str(values[1]),
                "verbatim_comment": "" if values[2] is None else str(values[2]),
                "comment_category": "" if values[3] is None else str(values[3]),
            }
        )
    return parsed


@app.on_event("startup")
def startup_event() -> None:
    init_db()
    ensure_default_manuscript()


@app.get("/", response_class=HTMLResponse)
def home(request: Request, manuscript_id: int | None = None) -> HTMLResponse:
    with get_conn() as conn:
        manuscripts = [dict(row) for row in conn.execute("SELECT id, name FROM manuscripts ORDER BY created_at, id")]
    if not manuscripts:
        default_id = ensure_default_manuscript()
        manuscripts = [{"id": default_id, "name": "Default Manuscript"}]

    selected_id = manuscript_id if manuscript_id else manuscripts[0]["id"]
    if selected_id not in [m["id"] for m in manuscripts]:
        selected_id = manuscripts[0]["id"]

    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "manuscripts": manuscripts,
            "selected_manuscript_id": selected_id,
            "next_open_behavior": "Next Open Ticket respects current search and filters.",
        },
    )


@app.get("/api/manuscripts")
def list_manuscripts() -> dict[str, Any]:
    with get_conn() as conn:
        rows = conn.execute("SELECT id, name, created_at FROM manuscripts ORDER BY created_at, id").fetchall()
    return {"manuscripts": [dict(row) for row in rows]}


@app.post("/api/manuscripts")
def create_manuscript(payload: dict[str, str]) -> dict[str, Any]:
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Manuscript name is required")
    with get_conn() as conn:
        try:
            cur = conn.execute("INSERT INTO manuscripts (name) VALUES (?)", (name,))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Could not create manuscript: {exc}") from exc
        manuscript = conn.execute(
            "SELECT id, name, created_at FROM manuscripts WHERE id = ?", (cur.lastrowid,)
        ).fetchone()
    return {"manuscript": dict(manuscript)}


@app.patch("/api/manuscripts/{manuscript_id}")
def rename_manuscript(manuscript_id: int, payload: dict[str, str]) -> dict[str, Any]:
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="New manuscript name is required")
    with get_conn() as conn:
        cur = conn.execute("UPDATE manuscripts SET name = ? WHERE id = ?", (name, manuscript_id))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail="Manuscript not found")
        manuscript = conn.execute(
            "SELECT id, name, created_at FROM manuscripts WHERE id = ?", (manuscript_id,)
        ).fetchone()
    return {"manuscript": dict(manuscript)}


@app.get("/api/manuscripts/{manuscript_id}/tickets")
def list_tickets(
    manuscript_id: int,
    search: str | None = Query(default=None),
    reviewer_id: str | None = Query(default=None),
    comment_category: str | None = Query(default=None),
    status: str | None = Query(default=None),
) -> dict[str, Any]:
    ensure_manuscript_exists(manuscript_id)
    tickets = query_tickets(manuscript_id, search, reviewer_id, comment_category, status)
    filter_values = get_filter_values(manuscript_id)
    return {"tickets": tickets, "filters": filter_values}


@app.post("/api/manuscripts/{manuscript_id}/import")
async def import_tickets(manuscript_id: int, file: UploadFile = File(...)) -> dict[str, Any]:
    ensure_manuscript_exists(manuscript_id)
    filename = file.filename or ""
    ext = Path(filename).suffix.lower()
    content = await file.read()

    if ext == ".csv":
        raw_rows = parse_csv_upload(content)
    elif ext == ".xlsx":
        raw_rows = parse_xlsx_upload(content)
    else:
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx imports are supported")

    prepared_rows = []
    for i, row in enumerate(raw_rows, start=1):
        try:
            prepared_rows.append(normalize_ticket_row(row))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Import row {i}: {exc}") from exc

    with get_conn() as conn:
        for row in prepared_rows:
            conn.execute(
                """
                INSERT INTO tickets (
                    manuscript_id,
                    reviewer_id,
                    reviewer_group_sort,
                    reviewer_num_sort,
                    line_number_display,
                    line_number_sort,
                    verbatim_comment,
                    comment_category,
                    response_text,
                    status,
                    created_at,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', 'OPEN', ?, ?)
                """,
                (
                    manuscript_id,
                    row["reviewer_id"],
                    row["reviewer_group_sort"],
                    row["reviewer_num_sort"],
                    row["line_number_display"],
                    row["line_number_sort"],
                    row["verbatim_comment"],
                    row["comment_category"],
                    now_iso(),
                    now_iso(),
                ),
            )

    return {"imported": len(prepared_rows)}


@app.get("/api/tickets/{ticket_id}")
def get_ticket(ticket_id: int) -> dict[str, Any]:
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Ticket not found")
    return {"ticket": dict(row)}


@app.patch("/api/tickets/{ticket_id}")
def update_ticket(ticket_id: int, payload: dict[str, Any]) -> dict[str, Any]:
    with get_conn() as conn:
        existing = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Ticket not found")

        updated = dict(existing)
        for field in [
            "reviewer_id",
            "line_number_display",
            "verbatim_comment",
            "comment_category",
            "response_text",
            "status",
        ]:
            if field in payload:
                updated[field] = payload[field]

        reviewer_id = str(updated["reviewer_id"]).strip()
        if not reviewer_id:
            raise HTTPException(status_code=400, detail="reviewer_id is required")

        line_display, line_sort = parse_line_number(updated["line_number_display"])
        if not line_display:
            raise HTTPException(status_code=400, detail="line_number is required")

        verbatim_comment = str(updated["verbatim_comment"]).strip()
        if not verbatim_comment:
            raise HTTPException(status_code=400, detail="verbatim_comment is required")

        category = str(updated["comment_category"]).strip().lower()
        if category not in VALID_CATEGORIES:
            raise HTTPException(status_code=400, detail="comment_category must be editorial, major, or minor")

        response_text = str(updated["response_text"])
        status = str(updated.get("status", "OPEN")).upper()
        if status not in {"OPEN", "COMPLETED"}:
            raise HTTPException(status_code=400, detail="status must be OPEN or COMPLETED")
        if status == "COMPLETED" and not response_text.strip():
            raise HTTPException(
                status_code=400,
                detail="response_text is required before marking a ticket completed",
            )

        reviewer_group_sort, reviewer_num_sort = parse_reviewer_sort(reviewer_id)

        conn.execute(
            """
            UPDATE tickets
            SET reviewer_id = ?,
                reviewer_group_sort = ?,
                reviewer_num_sort = ?,
                line_number_display = ?,
                line_number_sort = ?,
                verbatim_comment = ?,
                comment_category = ?,
                response_text = ?,
                status = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                reviewer_id,
                reviewer_group_sort,
                reviewer_num_sort,
                line_display,
                line_sort,
                verbatim_comment,
                category,
                response_text,
                status,
                now_iso(),
                ticket_id,
            ),
        )
        ticket = conn.execute("SELECT * FROM tickets WHERE id = ?", (ticket_id,)).fetchone()

    return {"ticket": dict(ticket)}


@app.get("/api/manuscripts/{manuscript_id}/next-open")
def next_open_ticket(
    manuscript_id: int,
    current_ticket_id: int | None = None,
    direction: str = Query(default="next"),
    search: str | None = Query(default=None),
    reviewer_id: str | None = Query(default=None),
    comment_category: str | None = Query(default=None),
) -> JSONResponse:
    ensure_manuscript_exists(manuscript_id)
    if direction not in {"next", "prev"}:
        raise HTTPException(status_code=400, detail="direction must be next or prev")

    open_tickets = query_tickets(
        manuscript_id,
        search=search,
        reviewer_id=reviewer_id,
        comment_category=comment_category,
        status="OPEN",
    )

    if not open_tickets:
        return JSONResponse({"ticket": None, "message": "No open tickets."})

    ids = [ticket["id"] for ticket in open_tickets]
    if current_ticket_id is None:
        chosen = open_tickets[0] if direction == "next" else open_tickets[-1]
        return JSONResponse({"ticket": chosen})

    with get_conn() as conn:
        current_row = conn.execute(
            "SELECT manuscript_id, reviewer_group_sort, reviewer_num_sort, line_number_sort, created_at, id FROM tickets WHERE id = ?",
            (current_ticket_id,),
        ).fetchone()

    if not current_row or current_row["manuscript_id"] != manuscript_id:
        chosen = open_tickets[0] if direction == "next" else open_tickets[-1]
        return JSONResponse({"ticket": chosen})

    if current_ticket_id in ids:
        idx = ids.index(current_ticket_id)
        if direction == "next":
            if idx + 1 < len(open_tickets):
                return JSONResponse({"ticket": open_tickets[idx + 1]})
            return JSONResponse({"ticket": open_tickets[0]})
        if idx - 1 >= 0:
            return JSONResponse({"ticket": open_tickets[idx - 1]})
        return JSONResponse({"ticket": open_tickets[-1]})

    current_tuple = (
        current_row["reviewer_group_sort"],
        current_row["reviewer_num_sort"],
        current_row["line_number_sort"],
        current_row["created_at"],
        current_row["id"],
    )

    if direction == "next":
        for ticket in open_tickets:
            ticket_tuple = (
                ticket["reviewer_group_sort"],
                ticket["reviewer_num_sort"],
                ticket["line_number_sort"],
                ticket["created_at"],
                ticket["id"],
            )
            if ticket_tuple > current_tuple:
                return JSONResponse({"ticket": ticket})
        return JSONResponse({"ticket": open_tickets[0]})

    for ticket in reversed(open_tickets):
        ticket_tuple = (
            ticket["reviewer_group_sort"],
            ticket["reviewer_num_sort"],
            ticket["line_number_sort"],
            ticket["created_at"],
            ticket["id"],
        )
        if ticket_tuple < current_tuple:
            return JSONResponse({"ticket": ticket})
    return JSONResponse({"ticket": open_tickets[-1]})
